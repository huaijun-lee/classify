# -*- coding: utf-8 -*-
"""
Created on Mon Jan  4 10:43:52 2016

@author: huaijun
"""
from pandas import read_excel 
import re
from openpyxl import Workbook
import time
import os
class Weibo:
    def __init__(self,excel_worksheet,source,defatul_f,log_f):
        self.worksheet = excel_worksheet
        self.col_name = list(self.worksheet.columns)
        self.rowname =['类别', '来源','帐号名称','链接','粉丝数（万）','更新时间','博主转发','防屏蔽转发','博主直发','防屏蔽直发','微任务转发','微任务直发','地域','平均转评','平均点赞','博主QQ号','博主微信','博主手机','描述','使用备注']
        self.source = source
        self.defatul_f = defatul_f
        self.log = log_f
    def location(self):
        n = len(self.col_name)
        loca = []
        test = ['称|号','链|址','粉|数']
        for i in test:       
            for j in range(n):                
                if re.search(i,self.col_name[j],re.I):
                    loca.append(j)                   
        for i in [r'转',r'直']:
            for j in range(n):
                if re.search(i,self.col_name[j]) and not re.search(r'硬|防|任',self.col_name[j]):
                    loca.append(j)                     
        return loca
    def deal_name(self,name):
        '''
        删除微博名称后面的V字
        '''
        name = str(name)
        name = re.sub('^ | $|v$','',name,re.I)
        return name
    def deal_fans(self,num):
        '''
        处理粉丝数的常见情况
        '''
        if not num == None:
            num = str(num)
            num = re.sub('\+$| ','',num)      
            num = re.sub('w$|万$|W$','',num)       
            if re.search(r'K|千',num,re.I):
                num = float(num[:-1]) / 10
            if float(num) > 10000:
                num = float(num) / 10000
            return num
    def deal_price(self,price):
        '''
        处理价格的常见情况
        '''
        if not price == None:
            price = str(price)
            price = re.sub('元$| ','',price)
            if re.search(r'万|W',price,re.I):
                price = float(price[:-1]) * 10000
            elif re.search(r'K|千',price,re.I):
                price = float(price[:-1]) * 1000
            return price
    def write(self):
        book = Workbook()
        weibo = book.get_active_sheet()
        weibo.title = 'Weibo'  # 设置worksheet的标题
        loca = self.location()
        for i in range(len(self.rowname)):
            weibo.cell(row = 1 , column = i + 1).value = self.rowname[i] 
        for i in range(len(self.worksheet.index)):
            try:            
                weibo.cell(row = i+2, column = 2).value = self.source   #写入来源
                weibo.cell(row = i+2, column = 3).value = self.deal_name(self.worksheet.values[i,loca[0]])  #写账号名称
                weibo.cell(row = i+2, column = 4).value = self.worksheet.values[i,loca[1]]  #写链接
                #if not re.search(r'保|不',str(self.worksheet.values[i,loca[2]])):
                #    weibo.cell(row = i+2, column = 5).value = self.deal_fans(self.worksheet.values[i,loca[2]])   #写粉丝数        
                price_bool =True
                if re.search(r'看|漫|谈|价|费|创|条',str(self.worksheet.values[i,loca[3]])):
                    weibo.cell(row = i + 2, column = 20).value = self.worksheet.values[i, loca[3]]
                    price_bool = False
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca[3]])):
                    price_bool = False
                if price_bool:
                    price_num = self.deal_price(self.worksheet.values[i,loca[3]])
                    if not price_num == 'nan':
                        weibo.cell(row = i+2, column = 7).value = price_num  #写转发价格
                second_bool = True
                if re.search(r'看|条|谈|费|漫|创|价',str(self.worksheet.values[i,loca[4]])):
                    weibo.cell(row = i + 2, column = 20).value = self.worksheet.values[i, loca[4]]
                    second_bool = False
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca[4]])):
                    second_bool = False
                if second_bool:
                    second_price = self.deal_price(self.worksheet.values[i,loca[4]])
                    if not second_price == 'nan':
                        weibo.cell(row = i+2, column = 9).value = second_price  #写直发价格
            except ValueError:
                self.log.write('处理数据时，第'+str(i+2)+'行数值格式有错\n')
            finally:
                pass
        loca_zhuan = 0  #硬广转发
        loca_zhi = 0    #硬广直发
        zhuan_bool = False  #硬广转
        zhi_bool = False    #硬广直
        for i in range(len(self.col_name)):
            if re.search(r'硬|防|任',self.col_name[i]) and re.search(r'转',self.col_name[i]):
                loca_zhuan = i
                zhuan_bool = True
            if re.search(r'硬|防|任',self.col_name[i]) and re.search(r'直',self.col_name[i]):
                loca_zhi = i
                zhi_bool = True
        if zhuan_bool :
            for i in range(len(self.worksheet.index)):
                price_bool =True
                if re.search(r'看|条|费|创|漫|价',str(self.worksheet.values[i,loca_zhuan])):
                    weibo.cell(row = i + 2, column = 20).value = self.worksheet.values[i, loca_zhuan]
                    price_bool = False
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca_zhuan])) :
                    price_bool = False
                if price_bool:
                    price_num = self.deal_price(self.worksheet.values[i,loca_zhuan])
                    if not price_num == 'nan':
                        weibo.cell(row = i+2, column = 8).value = price_num  #写防屏蔽转发价格
        if zhi_bool:
            for i in range(len(self.worksheet.index)):
                second_bool = True
                if re.search(r'看|条|费|创|漫|价',str(self.worksheet.values[i,loca_zhi])):
                    weibo.cell(row = i + 2, column = 20).value = self.worksheet.values[i, loca_zhi]
                    second_bool = False
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca_zhi])):
                    second_bool = False
                if second_bool:
                    second_price = self.deal_price(self.worksheet.values[i,loca_zhi])
                    if not second_price == 'nan':
                        weibo.cell(row = i+2, column = 10).value = second_price  #写防屏蔽直发价格
        self.check(weibo)      
        f_date = str(time.strftime("%y-%m-%d", time.localtime()))
        f_type = r'Weibo' 
        filepath = self.defatul_f +f_date+ f_type + '-' + self.source + '.xlsx'
        book.save(filepath)
        self.log.write('\n处理完成\n\t^_^ \n')
    def check(self,weibo):
        for i in range(len(self.worksheet.index)):
            try:                            
                if weibo.cell(row= i+2,column = 7).value != None and weibo.cell(row = i + 2, column = 9).value != None :
                    if float(weibo.cell(row= i+2,column = 7).value) > float(weibo.cell(row = i + 2, column = 9).value):
                        k = float(weibo.cell(row= i+2,column = 7).value)
                        weibo.cell(row= i+2,column = 7).value = float(weibo.cell(row = i + 2, column = 9).value)
                        weibo.cell(row = i + 2, column = 9).value = k
                if weibo.cell(row= i+2,column = 8).value != None and weibo.cell(row = i + 2, column = 10).value != None :
                    if float(weibo.cell(row= i+2,column = 8).value) > float(weibo.cell(row = i + 2, column = 10).value):
                        h = float(weibo.cell(row= i+2,column = 8).value)
                        weibo.cell(row= i+2,column = 8).value = float(weibo.cell(row = i + 2, column = 10).value)
                        weibo.cell(row = i + 2, column = 10).value = h
                if weibo.cell(row= i+2,column = 7).value != None and weibo.cell(row = i + 2, column = 8).value != None :
                    if float(weibo.cell(row= i+2,column = 7).value) > float(weibo.cell(row = i + 2, column = 8).value):
                        k = float(weibo.cell(row= i+2,column = 7).value)
                        weibo.cell(row= i+2,column = 7).value = float(weibo.cell(row = i + 2, column = 8).value)
                        weibo.cell(row = i + 2, column = 8).value = k
                if weibo.cell(row= i+2,column = 9).value != None and weibo.cell(row = i + 2, column = 10).value != None :
                    if float(weibo.cell(row= i+2,column = 9).value) > float(weibo.cell(row = i + 2, column = 10).value):
                        h = float(weibo.cell(row= i+2,column = 9).value)
                        weibo.cell(row= i+2,column = 9).value = float(weibo.cell(row = i + 2, column = 10).value)
                        weibo.cell(row = i + 2, column = 10).value = h
            except ValueError:
                self.log.write('检查时，第'+str(i+2)+'行数值有错\n')
            finally:
                pass                   
class WeChat:
    def __init__(self,excel_worksheet,source,defatul_f,log_f):
        self.worksheet = excel_worksheet
        self.col_name = list(self.worksheet.columns)        
        self.source = source
        self.defatul_f = defatul_f
        self.log = log_f
        self.rowname =['类型', '来源',	'微信号',	'微信ID',	'粉丝数（万）',	'更新时间',	'头条软文',	'次条软文',	'头条硬广',	'次条硬广',	'头条阅读',	'头条点赞',	'博主QQ号',	'博主微信',	'博主手机',	'备注']
    def location(self):
        '''
        匹配文件中有用数据的位置
        '''
        n = len(self.col_name)
        loca = []
        test = ['名|称','号|ID','粉|数|量']
        for i in test:       
            for j in range(n):            
                find=re.search(i , self.col_name[j],re.I)
                if find:
                    loca.append(j)
        for i in [r'头|首',r'次|2|二']:
            for j in range(n):
                if re.search(i,self.col_name[j]) and not re.search(r'硬',self.col_name[j]):
                    loca.append(j)
        return loca
    def deal_id(self,id):
        id = str(id)
        id = re.sub(' |\u200b','',id)
        return id
    def deal_fans(self,num):
        '''
        处理粉丝数的常见情况
        '''
        if not num == None:
            num = str(num)
            num = re.sub('\+$| ','',num)      
            num = re.sub('w$|万$|W$','',num)       
            if re.search(r'K|千',num,re.I):
                num = float(num[:-1]) / 10
            if float(num) > 5000:
                num = float(num) / 10000
            return num
    def deal_price(self,price):
        '''
        处理价格的常见情况
        '''
        if not price == None:
            price = str(price)
            price = re.sub('元$| ','',price)
            if re.search(r'万|W',price,re.I):
                price = float(price[:-1]) * 10000
            elif re.search(r'K|千',price,re.I):
                price = float(price[:-1]) * 1000
            return price
    def write(self):
        book = Workbook()
        wechat = book.get_active_sheet()
        wechat.title = 'WeChat'  # 设置worksheet的标题
        loca = self.location() 
        for i in range(len(self.rowname)):
        	wechat.cell(row = 1 , column = i + 1).value = self.rowname[i]
        for i in range(len(self.worksheet.index)):
            try:
                wechat.cell(row = i+2, column = 2).value = self.source   #写入来源
                wechat.cell(row = i+2, column = 3).value = self.deal_id(self.worksheet.values[i,loca[0]])  #写微信号
                wechat.cell(row = i+2, column = 4).value = self.worksheet.values[i,loca[1]]  #写微信ID
                if not re.search(r'保|不',str(self.worksheet.values[i,loca[2]])):
                    wechat.cell(row = i+2, column = 5).value = self.deal_fans(self.worksheet.values[i,loca[2]])   #写粉丝数        
                price_bool =True 
                if re.search(r'看|条|谈|费|漫|价|创|专',str(self.worksheet.values[i,loca[3]])):
                    wechat.cell(row = i + 2, column = 16).value = self.worksheet.values[i, loca[3]]
                    price_bool = False
                if re.search(r'不|拒',str(self.worksheet.values[i,loca[3]])):
                    wechat.cell(row = i + 2, column = 16).value = r'头条软文' + self.worksheet.values[i,loca[3]]
                    price_bool = False
                if re.search(r'\/|\\',str(self.worksheet.values[i,loca[3]])):
                    wechat.cell(row = i + 2, column = 16).value = r'不接头条软文'
                    price_bool = False
                if price_bool:
                    first_price = self.deal_price(self.worksheet.values[i,loca[3]])
                    if not first_price=='nan':
                        wechat.cell(row = i+2, column = 7).value = first_price  #写头条软文的价格
                second_bool = True
                if re.search(r'看|条|谈|费|创|漫|价',str(self.worksheet.values[i,loca[4]])):
                    wechat.cell(row = i + 2, column = 16).value = self.worksheet.values[i, loca[4]]
                    second_bool = False
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca[4]])):
                    second_bool = False
                if second_bool:
                    second_price = self.deal_price(self.worksheet.values[i,loca[4]])
                    if not second_price == 'nan':
                        wechat.cell(row = i+2, column = 8).value = second_price  #写次条软文的价格
            except ValueError:
                self.log.write('处理数据时，第'+str(i+2)+'行数值有错\n')
            finally:
                pass  
        loca_first = 0      #硬广头条
        loca_second = 0     #硬广次条
        first_bool = False  #硬广头
        second_bool = False    #硬广次
        for i in range(len(self.col_name)):
            if re.search(r'硬',self.col_name[i]) and re.search(r'头',self.col_name[i]):
                loca_first = i
                first_bool = True
            if re.search(r'硬',self.col_name[i]) and re.search(r'次|2|二',self.col_name[i]):
                loca_second = i
                second_bool = True
        if first_bool :
            for i in range(len(self.worksheet.index)):
                price_bool =True
                if re.search(r'看|条|费|漫|创|价',str(self.worksheet.values[i,loca_first])):
                    wechat.cell(row = i + 2, column = 16).value = self.worksheet.values[i, loca_first]
                    price_bool = False
                    pass
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca_first])) :
                    price_bool = False
                    pass
                if price_bool:
                    price_num = self.deal_price(self.worksheet.values[i,loca_first])
                    if not price_num == 'nan':
                        wechat.cell(row = i+2, column = 9).value = price_num  #写硬广头条价格
        if second_bool:    
            for i in range(len(self.worksheet.index)):
                second_bool = True
                if re.search(r'看|条|费|创|漫|价',str(self.worksheet.values[i,loca_second])):
                    wechat.cell(row = i + 2, column = 16).value = self.worksheet.values[i, loca_second]
                    second_bool = False
                    pass
                if re.search(r'不|拒|\/',str(self.worksheet.values[i,loca_second])):
                    second_bool = False
                if second_bool:
                    second_price = self.deal_price(self.worksheet.values[i,loca_second])
                    if not second_price == 'nan':
                        wechat.cell(row = i+2, column = 10).value = second_price  #写硬广次条价格
        self.check(wechat)        
        f_date = str(time.strftime("%y-%m-%d", time.localtime()))
        f_type = r'WeChat' 
        filepath = self.defatul_f +f_date+ f_type + '-' + self.source + '.xlsx'
        book.save(filepath)
        self.log.write('\n处理完成\n\t^_^ \n')
    def check(self,wechat):
        for i in range(len(self.worksheet.index)):
            try:      
                if wechat.cell(row= i+2,column = 7).value != None and wechat.cell(row = i + 2, column = 8).value != None :
                    if float(wechat.cell(row= i+2,column = 8).value) > float(wechat.cell(row = i + 2, column = 7).value):
                        k = float(wechat.cell(row= i+2,column = 8).value)
                        wechat.cell(row= i+2,column = 8).value = float(wechat.cell(row = i + 2, column = 7).value)
                        wechat.cell(row = i + 2, column = 7).value = k
                if wechat.cell(row= i+2,column = 9).value != None and wechat.cell(row = i + 2, column = 10).value != None :
                    if float(wechat.cell(row= i+2,column = 10).value) > float(wechat.cell(row = i + 2, column = 9).value):
                        h = float(wechat.cell(row= i+2,column = 10).value)
                        wechat.cell(row= i+2,column = 10).value = float(wechat.cell(row = i + 2, column = 9).value)
                        wechat.cell(row = i + 2, column = 9).value = h
                if wechat.cell(row= i+2,column = 7).value != None and wechat.cell(row = i + 2, column = 9).value != None :
                    if float(wechat.cell(row= i+2,column = 7).value) > float(wechat.cell(row = i + 2, column = 9).value):
                        k = float(wechat.cell(row= i+2,column = 7).value)
                        wechat.cell(row= i+2,column = 7).value = float(wechat.cell(row = i + 2, column = 9).value)
                        wechat.cell(row = i + 2, column = 9).value = k
                if wechat.cell(row= i+2,column = 8).value != None and wechat.cell(row = i + 2, column = 10).value != None :
                    if float(wechat.cell(row= i+2,column = 8).value) > float(wechat.cell(row = i + 2, column = 10).value):
                        h = float(wechat.cell(row= i+2,column = 8).value)
                        wechat.cell(row= i+2,column = 8).value = float(wechat.cell(row = i + 2, column = 10).value)
                        wechat.cell(row = i + 2, column = 10).value = h
            except ValueError:
                self.log.write('检查时，第'+str(i+2)+'行数值有错\n')
            finally:
                pass      
if __name__ == '__main__':
    now_f_url = os.getcwd()
    inputfile_url = now_f_url +'\\待处理'
    outputfile_url = now_f_url +'\\已处理\\'
    list_dirs = os.listdir(inputfile_url)
    if not os.path.exists(outputfile_url):
        os.makedirs(outputfile_url)
    log_txt = './log'
    log_date = str(time.strftime("%y-%m-%d-%H-%M", time.localtime()))
    log_txt_url = log_txt+log_date+'.txt'
    with open(log_txt_url,'a+') as log_f: 
        for files in list_dirs:
            fileurl = os.path.join(inputfile_url,files)
            log_f.write('处理'+files+'文件\n')
            try:       
                files =re.sub('.xlsx$','',files)
                files_split = files.split('_',2)
                source = files_split[0]
                file_type = files_split[1]
                excel_worksheet = read_excel(fileurl)
                excel_colname = list(excel_worksheet.columns)
                if re.search(r'微博',file_type):
                    file_weibo = Weibo(excel_worksheet,source,outputfile_url,log_f)
                    file_weibo.write()
                elif re.search(r'微信',file_type):
                    file_wechat = WeChat(excel_worksheet,source,outputfile_url,log_f)
                    file_wechat.write()
                else:
                    log_f.write('文件\t' + files +'\t中有错误，无法识别该文件\n')
            except IndexError:
                log_f.write(files+'文件名格式有误\n')
